import re
from pathlib import Path

import openpyxl

try:
    import xlrd
except ImportError:
    xlrd = None

from ..app_types import StudentData


def process_data() -> None:
    print("=== Starting data processing ===")

    root_dir = Path(__file__).resolve().parent.parent
    score_template_dir = root_dir / "score-templates"
    comment_template_dir = root_dir / "comment-templates"
    input_dir = root_dir / "inputs"

    print(f"Score templates: {score_template_dir}")
    print(f"Comment templates: {comment_template_dir}")
    print(f"Inputs: {input_dir}")

    _process_data_real(score_template_dir, comment_template_dir, input_dir)
    print("=== Data processing completed ===\\n")


def _normalize_subject_name(subject_name: str) -> str:
    return re.sub(r"_(hn|ht)$", "", subject_name)


def _parse_sheet_name(sheet_name: str) -> tuple[str, str]:
    parts = sheet_name.split("_")

    if len(parts) >= 3:
        last = parts[-1]
        second_last = parts[-2]
        if second_last.isdigit() and last.isdigit():
            class_name = f"{second_last}_{last}"
            subject = "_".join(parts[:-2])
        else:
            class_name = last
            subject = "_".join(parts[:-1])
    else:
        class_name = parts[-1]
        subject = "_".join(parts[:-1])

    return _normalize_subject_name(subject), class_name


def _convert_xls_to_xlsx(xls_path: Path) -> Path:
    if xlrd is None:
        raise RuntimeError("xlrd is required to convert .xls files")

    xls_workbook = xlrd.open_workbook(str(xls_path))
    xlsx_workbook = openpyxl.Workbook()
    default_sheet = xlsx_workbook.active
    xlsx_workbook.remove(default_sheet)

    for sheet_name in xls_workbook.sheet_names():
        sheet = xls_workbook.sheet_by_name(sheet_name)
        ws = xlsx_workbook.create_sheet(title=sheet_name)
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=sheet.cell_value(row_idx, col_idx))

    xlsx_path = xls_path.with_suffix(".xlsx")
    xlsx_workbook.save(xlsx_path)
    return xlsx_path


def _read_input_file(file_path: Path) -> dict[str, list[StudentData]]:
    data: dict[str, list[StudentData]] = {}

    working_file = file_path
    if file_path.suffix.lower() == ".xls":
        print("Converting .xls to .xlsx...")
        working_file = _convert_xls_to_xlsx(file_path)

    workbook = openpyxl.load_workbook(working_file, data_only=True)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        subject, class_name = _parse_sheet_name(sheet_name)
        key = f"{subject}_{class_name}"
        students: list[StudentData] = []

        for row in worksheet.iter_rows(min_row=1, values_only=True):
            cell_a = row[0] if len(row) > 0 else None
            is_student_row = isinstance(cell_a, int) or (
                isinstance(cell_a, float) and cell_a.is_integer()
            ) or (isinstance(cell_a, str) and cell_a.isdigit())

            if not is_student_row:
                continue

            name_part1 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            name_part2 = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            student_name = f"{name_part1} {name_part2}".strip()

            frequency_scores: list[float] = []
            for idx in [5, 6, 7, 8]:
                if len(row) > idx and row[idx] not in (None, ""):
                    frequency_scores.append(float(row[idx]))

            midterm_score = float(row[9]) if len(row) > 9 and row[9] not in (None, "") else None
            finalterm_score = float(row[10]) if len(row) > 10 and row[10] not in (None, "") else None
            comment = str(row[12]).strip() if len(row) > 12 and row[12] not in (None, "") else ""

            students.append(
                StudentData(
                    name=student_name,
                    frequency_scores=frequency_scores,
                    midterm_score=midterm_score,
                    finalterm_score=finalterm_score,
                    comment=comment,
                )
            )

        if students:
            data[key] = students
            print(f"Read {len(students)} students from sheet '{sheet_name}' ({key})")

    return data


def _find_output_files(score_dir: Path, comment_dir: Path, subject: str, class_name: str) -> tuple[list[Path], list[Path]]:
    score_files = list(score_dir.glob(f"diem_hoc_ky_*_lop_{class_name}_mon_{subject}.xlsx"))
    class_no_underscore = class_name.replace("_", "")

    comment_files = set(comment_dir.glob(f"nhan_xet_hoc_ky_*_lop_{class_no_underscore}_mon_{subject}.xlsx"))
    comment_files.update(comment_dir.glob(f"nhan_xet_hoc_ky_*_lop_{class_name}_mon_{subject}.xlsx"))

    return score_files, list(comment_files)


def _fill_score_file(file_path: Path, students: list[StudentData]) -> None:
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[workbook.sheetnames[0]]
    student_map = {s.name: s for s in students}
    filled_count = 0

    for row in worksheet.iter_rows(min_row=1):
        cell_a = row[0].value
        is_student_row = isinstance(cell_a, int) or (
            isinstance(cell_a, float) and cell_a.is_integer()
        ) or (isinstance(cell_a, str) and cell_a.isdigit())

        if not is_student_row:
            continue

        student_name = str(row[2].value).strip() if row[2].value else ""
        student = student_map.get(student_name)
        if not student:
            continue

        if student.midterm_score is not None:
            row[9].value = student.midterm_score
        if student.finalterm_score is not None:
            row[10].value = student.finalterm_score

        for idx, score in enumerate(student.frequency_scores):
            target_col = 4 + idx
            if target_col < len(row):
                row[target_col].value = score

        filled_count += 1

    workbook.save(file_path)
    print(f"  Filled {filled_count} students in score file: {file_path.name}")


def _fill_comment_file(file_path: Path, students: list[StudentData]) -> None:
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[workbook.sheetnames[0]]
    student_map = {s.name: s for s in students}
    filled_count = 0

    for row in worksheet.iter_rows(min_row=1):
        cell_a = row[0].value
        is_student_row = isinstance(cell_a, int) or (
            isinstance(cell_a, float) and cell_a.is_integer()
        ) or (isinstance(cell_a, str) and cell_a.isdigit())

        if not is_student_row:
            continue

        student_name = str(row[2].value).strip() if row[2].value else ""
        student = student_map.get(student_name)
        if student and student.comment:
            row[6].value = student.comment
            filled_count += 1

    workbook.save(file_path)
    print(f"  Filled {filled_count} students in comment file: {file_path.name}")


def _process_data_real(score_template_dir: Path, comment_template_dir: Path, input_dir: Path) -> None:
    input_file = input_dir / "so_diem_cac_mon_toi_day.xls"
    if not input_file.exists():
        alt_xlsx = input_dir / "so_diem_cac_mon_toi_day.xlsx"
        if alt_xlsx.exists():
            input_file = alt_xlsx
        else:
            print(f"❌ Input file not found: {input_file}")
            return

    print("\\nReading input file...")
    student_data = _read_input_file(input_file)

    print(f"\\nFound {len(student_data)} subject-class combinations")
    print("\\nProcessing output files...")

    for key, students in student_data.items():
        parts = key.split("_")
        class_name = f"{parts[-2]}_{parts[-1]}"
        subject = "_".join(parts[:-2])

        print(f"\\nSubject: {subject}, Class: {class_name}")
        score_files, comment_files = _find_output_files(score_template_dir, comment_template_dir, subject, class_name)

        if score_files:
            for score_file in score_files:
                _fill_score_file(score_file, students)
        else:
            print(f"  ⚠ No score template found for {subject}_{class_name}")

        if comment_files:
            for comment_file in comment_files:
                _fill_comment_file(comment_file, students)
        else:
            print(f"  ⚠ No comment template found for {subject}_{class_name}")

    print("\\n✓ Processing complete!")
